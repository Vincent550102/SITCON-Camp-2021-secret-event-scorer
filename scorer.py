import openpyxl
import json
from env import CHA1, CHA2, curruptcolor
# workbook = openpyxl.load_workbook('SITCON Camp Online 2021 挑戰賽一.xlsx')
workbook = openpyxl.load_workbook('SITCON Camp Online 2021 挑戰賽二.xlsx')
# nowcha = CHA1
nowcha = CHA2

sheet = workbook['Kahoot! Summary']
mxR = sheet.max_row

if __name__ == '__main__':
    result = dict()
    for r in range(4, mxR+1):
        score = 0
        for num in CHA1:
            score += CHA1[num] * \
                curruptcolor[str(sheet.cell(
                    row=r, column=2+2*num).fill.start_color.index)]
        result[sheet.cell(row=r, column=2).value] = score
    result = sorted(result.items(), key=lambda x: -x[1])
    rank = dict()
    rankcnt = 1
    i = 0
    while i < len(result):
        ranklist = list()
        ranklist.append(result[i])
        while i < len(result)-1 and result[i][1] == result[i+1][1]:
            i += 1
            ranklist.append(result[i])
            print(i)
        else:
            i += 1
        rank[f'Rank{str(rankcnt)}'] = ranklist
        rankcnt += 1

    # print(rank)
    with open('rank.json', 'w', encoding='utf-8') as f:
        json.dump(rank, f, ensure_ascii=False, indent=4)
